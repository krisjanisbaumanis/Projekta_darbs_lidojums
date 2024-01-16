import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

from openpyxl import Workbook, load_workbook
import time

import smtplib
from email.message import EmailMessage
from os.path import basename
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


service = Service()
option = webdriver.ChromeOptions()
option.add_argument('--headless')
driver = webdriver.Chrome(service=service, options= option)

url = "https://www.ryanair.com/lv/en/trip/flights/select?adults=2&teens=0&children=0&infants=0&dateOut=2024-05-13&dateIn=2024-05-18&isConnectedFlight=false&discount=0&promoCode=&isReturn=true&originIata=RIX&destinationIata=TPS&tpAdults=2&tpTeens=0&tpChildren=0&tpInfants=0&tpStartDate=2024-05-13&tpEndDate=2024-05-18&tpDiscount=0&tpPromoCode=&tpOriginIata=RIX&tpDestinationIata=TPS"

max_retries = 3
retries = 0

while retries<max_retries:
    try:
        driver.get(url)
        break
    except Exception as e:
        print(f"Error: {e}")
        retries = retries + 1

time.sleep(1)

find = driver.find_element(By.CLASS_NAME, "cookie-popup-with-overlay__button")
find.click()

time.sleep(1)

find.clear
sar = []
find = driver.find_element(By.CSS_SELECTOR, "body > app-root > flights-root > div > div > div > div > flights-lazy-content > flights-summary-container > flights-summary > div > div:nth-child(1) > journey-container > journey > div")
teksts_no = find.text.split('\n')[0]
teksts_par = find.text.split('\n')[11:12]
cena = find.text.split('\n')[13:17]
cena = ''.join(cena)
turp = [teksts_no] + teksts_par + [cena]
print(turp)


time.sleep(1)


find.clear 
sar.clear
find = driver.find_element(By.CSS_SELECTOR, "body > app-root > flights-root > div > div > div > div > flights-lazy-content > flights-summary-container > flights-summary > div > div.inbound.ng-tns-c143892162-4.inbound--unconfirmed > journey-container > journey")
teksts = find.text.split('\n')[0]
lid = find.text.split('\n')[8:14]
lidojums = find.text.split('\n')[8:9]
cena = lid[2:7]
cena = ''.join(cena)
atpakal = [teksts] + lidojums + [cena]
print(atpakal)

#exceli ierakstām izgūto inf

wb = load_workbook("Cenas.xlsx")
ws = wb["Sheet1"]
ws.append(turp)
ws.append(atpakal)
empt_row = [' ']*ws.max_column 
ws.append(empt_row) # ieliekam vienu tukšu rindu
wb.save("Cenas.xlsx")
wb.close()

from_adr = 'krishjaanis.baumanis@gmail.com'
to_adr = 'krishjaanis.baumanis@gmail.com'
subject = 'Info + Excelis'
text = f'{turp}\n{atpakal}'

msg = MIMEMultipart()
msg['From'] = from_adr
msg['To'] = to_adr
msg['Subject'] = subject
body = MIMEText(text, 'plain')
msg.attach(body)

password = "gbsmuhvxtshozcwy"
file = 'Cenas.xlsx'
with open(file, 'rb') as f:
    attachment = MIMEApplication(f.read(), Name=basename(file))
    attachment['Content-Disposition'] = 'attachment; filename="{}"'.format(basename(file))
msg.attach(attachment)
server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(from_adr, password)
server.send_message(msg, from_addr=from_adr, to_addrs=to_adr)
